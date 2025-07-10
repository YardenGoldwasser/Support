from flask import Flask, request, jsonify, render_template, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from io import BytesIO

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///support_bot.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
CORS(app)

# Database Models
class Issue(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=False)
    customer_name = db.Column(db.String(100), nullable=False)
    customer_email = db.Column(db.String(120), nullable=False)
    customer_phone = db.Column(db.String(20), nullable=True)
    issue_type = db.Column(db.String(50), nullable=False)  # 'bug', 'feature_request', 'support'
    priority = db.Column(db.String(20), default='medium')  # 'low', 'medium', 'high', 'critical'
    status = db.Column(db.String(20), default='open')  # 'open', 'in_progress', 'resolved', 'closed'
    assigned_to = db.Column(db.String(100), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    resolved_at = db.Column(db.DateTime, nullable=True)
    internal_notes = db.Column(db.Text, nullable=True)
    
    def to_dict(self):
        return {
            'id': self.id,
            'title': self.title,
            'description': self.description,
            'customer_name': self.customer_name,
            'customer_email': self.customer_email,
            'customer_phone': self.customer_phone,
            'issue_type': self.issue_type,
            'priority': self.priority,
            'status': self.status,
            'assigned_to': self.assigned_to,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None,
            'resolved_at': self.resolved_at.isoformat() if self.resolved_at else None,
            'internal_notes': self.internal_notes
        }

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/admin')
def admin():
    return render_template('admin.html')

@app.route('/api/issues', methods=['POST'])
def create_issue():
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['title', 'description', 'customer_name', 'customer_email', 'issue_type']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} is required'}), 400
        
        # Create new issue
        issue = Issue(
            title=data['title'],
            description=data['description'],
            customer_name=data['customer_name'],
            customer_email=data['customer_email'],
            customer_phone=data.get('customer_phone'),
            issue_type=data['issue_type'],
            priority=data.get('priority', 'medium')
        )
        
        db.session.add(issue)
        db.session.commit()
        
        return jsonify({
            'message': 'Issue created successfully',
            'issue_id': issue.id,
            'issue': issue.to_dict()
        }), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/issues', methods=['GET'])
def get_issues():
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        status = request.args.get('status')
        issue_type = request.args.get('issue_type')
        priority = request.args.get('priority')
        
        query = Issue.query
        
        # Apply filters
        if status:
            query = query.filter(Issue.status == status)
        if issue_type:
            query = query.filter(Issue.issue_type == issue_type)
        if priority:
            query = query.filter(Issue.priority == priority)
        
        # Order by creation date (newest first)
        query = query.order_by(Issue.created_at.desc())
        
        # Paginate
        issues = query.paginate(page=page, per_page=per_page, error_out=False)
        
        return jsonify({
            'issues': [issue.to_dict() for issue in issues.items],
            'total': issues.total,
            'pages': issues.pages,
            'current_page': page,
            'per_page': per_page
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/issues/<int:issue_id>', methods=['GET'])
def get_issue(issue_id):
    try:
        issue = Issue.query.get_or_404(issue_id)
        return jsonify(issue.to_dict())
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/issues/<int:issue_id>', methods=['PUT'])
def update_issue(issue_id):
    try:
        issue = Issue.query.get_or_404(issue_id)
        data = request.get_json()
        
        # Update allowed fields
        updatable_fields = ['status', 'priority', 'assigned_to', 'internal_notes']
        for field in updatable_fields:
            if field in data:
                setattr(issue, field, data[field])
        
        # Set resolved_at if status is resolved or closed
        if data.get('status') in ['resolved', 'closed'] and not issue.resolved_at:
            issue.resolved_at = datetime.utcnow()
        elif data.get('status') not in ['resolved', 'closed']:
            issue.resolved_at = None
        
        issue.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'message': 'Issue updated successfully',
            'issue': issue.to_dict()
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/issues/export', methods=['GET'])
def export_issues():
    try:
        # Get all issues
        issues = Issue.query.order_by(Issue.created_at.desc()).all()
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Support Issues"
        
        # Define headers
        headers = [
            'ID', 'Title', 'Description', 'Customer Name', 'Customer Email', 
            'Customer Phone', 'Issue Type', 'Priority', 'Status', 'Assigned To',
            'Created At', 'Updated At', 'Resolved At', 'Internal Notes'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for row, issue in enumerate(issues, 2):
            ws.cell(row=row, column=1, value=issue.id)
            ws.cell(row=row, column=2, value=issue.title)
            ws.cell(row=row, column=3, value=issue.description)
            ws.cell(row=row, column=4, value=issue.customer_name)
            ws.cell(row=row, column=5, value=issue.customer_email)
            ws.cell(row=row, column=6, value=issue.customer_phone)
            ws.cell(row=row, column=7, value=issue.issue_type)
            ws.cell(row=row, column=8, value=issue.priority)
            ws.cell(row=row, column=9, value=issue.status)
            ws.cell(row=row, column=10, value=issue.assigned_to)
            ws.cell(row=row, column=11, value=issue.created_at.strftime('%Y-%m-%d %H:%M:%S') if issue.created_at else '')
            ws.cell(row=row, column=12, value=issue.updated_at.strftime('%Y-%m-%d %H:%M:%S') if issue.updated_at else '')
            ws.cell(row=row, column=13, value=issue.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if issue.resolved_at else '')
            ws.cell(row=row, column=14, value=issue.internal_notes)
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'support_issues_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/stats', methods=['GET'])
def get_stats():
    try:
        total_issues = Issue.query.count()
        open_issues = Issue.query.filter(Issue.status == 'open').count()
        in_progress_issues = Issue.query.filter(Issue.status == 'in_progress').count()
        resolved_issues = Issue.query.filter(Issue.status.in_(['resolved', 'closed'])).count()
        
        # Issues by type
        bug_count = Issue.query.filter(Issue.issue_type == 'bug').count()
        feature_count = Issue.query.filter(Issue.issue_type == 'feature_request').count()
        support_count = Issue.query.filter(Issue.issue_type == 'support').count()
        
        # Issues by priority
        critical_count = Issue.query.filter(Issue.priority == 'critical').count()
        high_count = Issue.query.filter(Issue.priority == 'high').count()
        medium_count = Issue.query.filter(Issue.priority == 'medium').count()
        low_count = Issue.query.filter(Issue.priority == 'low').count()
        
        return jsonify({
            'total_issues': total_issues,
            'open_issues': open_issues,
            'in_progress_issues': in_progress_issues,
            'resolved_issues': resolved_issues,
            'by_type': {
                'bug': bug_count,
                'feature_request': feature_count,
                'support': support_count
            },
            'by_priority': {
                'critical': critical_count,
                'high': high_count,
                'medium': medium_count,
                'low': low_count
            }
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5000)